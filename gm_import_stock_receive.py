"""Convert GM_stock_list.xlsx into an Autocount stock-receive import file.

See CLAUDE.md and /home/marcus/.claude/plans/have-a-read-of-cheeky-newt.md for the
full spec this implements.
"""
import re
import json
import datetime
from collections import defaultdict

import yaml
import openpyxl

BASE = "/home/marcus/stock_receivables_import"

# Grade codes confirmed present both in GM description text and in imported_item_codes'
# UDF_Grade column -- used as a disambiguation signal (see resolve_records), not just
# discarded like a generic qualifier.
GRADE_WORDS = {"XF", "HG", "SG", "AG", "PG", "USXF", "WXP", "WXPR", "WXPB"}

# Generic qualifier words confirmed present in GM description text that are NOT
# brands and must never trigger the "unresolved brand" exclusion.
IGNORE_WORDS = {
    "WASH", "PREMIUM", "GENERIC", "SAMPLE", "FAMILY", "PACK", "LARGE", "SMALL",
    "TUBE", "OPEN", "TOP", "BLUSH", "STRIPE", "GB", "ORGANIC", "MIX",
} | GRADE_WORDS

TEMPLATE_COLUMNS = [
    "DocNo", "DocDate", "Description", "Note", "Remark1", "Remark2", "Remark3",
    "Remark4", "RefDocNo", "UDF_RCV_MArrivalDate", "Numbering", "ItemCode",
    "DetailDescription", "FurtherDescription", "Location", "ProjNo", "Qty", "UOM",
    "UnitCost", "SubTotal", "PrintOut", "UDF_RCVDtl_ETA", "UDF_RCVDtl_DnChar",
    "UDF_RCVDtl_Remarks", "UDF_RCVDtl_STPrice",
]

REVIEW_COLUMNS = [
    "DocNo", "Location", "DetailDescription", "Size", "Qty", "Status", "Reason",
    "ParsedFruit", "ParsedCountry", "UnmatchedText", "CandidateItemCodes", "SourceRow",
]


def tokenize(text):
    return [t.upper() for t in re.findall(r"[A-Za-z][A-Za-z0-9]*", text or "")]


def singularize(word):
    lw = word.lower()
    if lw.endswith("ies"):
        return word[:-3] + "y"
    if lw.endswith("oes"):
        return word[:-2]
    if lw.endswith("ses") or lw.endswith("xes"):
        return word[:-2]
    if lw.endswith("s") and not lw.endswith("ss"):
        return word[:-1]
    return word


def load_lookups():
    with open(f"{BASE}/item_code_hashmap.json") as f:
        hashmap = json.load(f)

    country_lookup = {}
    for k, v in hashmap["country"].items():
        country_lookup[k.upper()] = k
        country_lookup[str(v).upper()] = k
    # Manual overrides confirmed necessary against real GM text (abbreviations the
    # hashmap doesn't itself contain).
    country_lookup["AUST"] = "Australia"
    country_lookup["US"] = "USA"

    wb = openpyxl.load_workbook(f"{BASE}/imported_item_codes.xlsx", data_only=True)
    ws = wb["Sheet1"]
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers) if h}

    item_master = []
    seen = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = row[idx["ItemCode"]]
        if not code:
            continue
        itype = row[idx["ItemType"]]
        icat = row[idx["ItemCategory"]]
        icls = row[idx["ItemClass"]]
        ibrand = row[idx["ItemBrand"]] or "Nil"
        isize = row[idx["UDF_UOM_Size"]]
        iweight = row[idx["UDF_UOM_WeightPerCtn"]]
        igrade = row[idx["UDF_Grade"]]
        key = (code, itype, icat, icls, ibrand, str(isize))
        if key in seen:
            continue
        seen.add(key)
        item_master.append({
            "ItemCode": code, "ItemType": itype, "ItemCategory": icat,
            "ItemClass": icls, "ItemBrand": str(ibrand).strip(),
            "Size": isize, "Weight": iweight,
            "Grade": str(igrade).strip().upper() if igrade else None,
        })

    fruit_lookup = {}
    for f in {r["ItemType"] for r in item_master if r["ItemType"]}:
        fruit_lookup[f.upper()] = f
        fruit_lookup[singularize(f).upper()] = f

    brand_set = set()
    with open(f"{BASE}/unique_brands.yaml") as f:
        yml = yaml.safe_load(f)
    for b in yml or []:
        brand_set.add(str(b).strip().upper())
    for r in item_master:
        if r["ItemBrand"] and r["ItemBrand"] != "Nil":
            brand_set.add(r["ItemBrand"].upper())

    by_type_class = defaultdict(list)
    for r in item_master:
        by_type_class[(r["ItemType"], r["ItemClass"])].append(r)

    return {
        "country_lookup": country_lookup,
        "fruit_lookup": fruit_lookup,
        "brand_set": brand_set,
        "by_type_class": by_type_class,
    }


def parse_date(raw):
    if raw is None:
        return None
    if isinstance(raw, datetime.datetime):
        return raw
    s = str(raw).strip()
    for fmt in ("%d/%m/%y", "%d/%m/%Y"):
        try:
            return datetime.datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def parse_size(raw):
    if raw is None:
        return None, None
    if isinstance(raw, (int, float)):
        return int(raw), str(int(raw))
    s = str(raw).strip()
    m = re.match(r"^(\d+)", s)
    if m:
        return int(m.group(1)), s
    return None, s


def parse_duty(raw):
    if raw is None:
        return None
    s = str(raw).strip().upper()
    m = re.match(r"^RM\s*([\d.]+)$", s)
    return float(m.group(1)) if m else None


def parse_price(raw):
    if raw is None:
        return None
    s = str(raw).strip()
    m = re.match(r"^\$\s*([\d.]+)$", s)
    return float(m.group(1)) if m else None


def size_eq(candidate_size, size_num):
    try:
        return int(float(str(candidate_size).strip())) == int(size_num)
    except (TypeError, ValueError):
        return str(candidate_size).strip() == str(size_num)


def parse_gm_sheet(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"]
    records = []
    header_recent = []
    last_full = None

    for r in range(5, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        c = ws.cell(row=r, column=3).value
        d = ws.cell(row=r, column=4).value
        e = ws.cell(row=r, column=5).value
        f = ws.cell(row=r, column=6).value
        g = ws.cell(row=r, column=7).value
        h = ws.cell(row=r, column=8).value
        i = ws.cell(row=r, column=9).value
        j = ws.cell(row=r, column=10).value
        k = ws.cell(row=r, column=11).value

        other_filled = any(v is not None for v in (c, d, e, f, i, k))

        if a is None and not other_filled and g is None and h is None:
            continue  # blank separator

        if a is not None and not other_filled and g is None and h is None:
            header_recent.append(str(a).strip())
            header_recent = header_recent[-2:]
            continue

        if a is not None:
            last_full = {
                "desc_text": str(a).strip(), "kg_raw": c, "size_raw": d,
                "container": e, "date_raw": f, "qty": g, "remark": h,
                "duty_raw": i, "currency": j, "price_raw": k, "source_row": r,
                "header_context": " ".join(header_recent),
            }
            records.append(dict(last_full, row_type="full"))
        elif other_filled:
            if last_full is None:
                continue
            rec = {
                "desc_text": last_full["desc_text"], "kg_raw": c, "size_raw": d,
                "container": e, "date_raw": f, "qty": g, "remark": h,
                "duty_raw": i, "currency": j, "price_raw": k, "source_row": r,
                "header_context": last_full["header_context"],
            }
            last_full = rec
            records.append(dict(rec, row_type="type1"))
        else:
            if last_full is None:
                continue
            rec = dict(last_full)
            rec["qty"] = g
            rec["remark"] = h
            rec["source_row"] = r
            records.append(dict(rec, row_type="type2"))

    return records


def make_flag(rec, status, reason, fruit=None, country=None, unmatched="", candidates=None):
    size_num, size_str = parse_size(rec.get("size_raw"))
    kg_display = str(rec["kg_raw"]).strip() if rec.get("kg_raw") is not None else ""
    detail_desc = f"{rec['desc_text']} {kg_display} {size_str or ''}".strip()
    qty = rec.get("qty")
    try:
        qty = int(float(qty)) if qty is not None else 0
    except (TypeError, ValueError):
        qty = 0
    return {
        "DocNo": None,
        "Location": str(rec.get("container") or "").strip(),
        "DetailDescription": detail_desc,
        "Size": size_num if size_num is not None else rec.get("size_raw"),
        "Qty": qty,
        "Status": status,
        "Reason": reason,
        "ParsedFruit": fruit,
        "ParsedCountry": country,
        "UnmatchedText": unmatched,
        "CandidateItemCodes": "; ".join(sorted({c["ItemCode"] for c in candidates})) if candidates else None,
        "SourceRow": rec.get("source_row"),
    }


def resolve_records(records, lookups):
    resolved = []
    flagged = []
    sticky_country = None
    sticky_fruit = None

    for rec in records:
        qty_raw = rec.get("qty")
        try:
            qty_num = float(qty_raw) if qty_raw is not None else 0
        except (TypeError, ValueError):
            qty_num = 0
        if qty_num <= 0:
            continue

        desc_tokens = tokenize(rec["desc_text"])
        header_tokens = tokenize(rec.get("header_context", ""))

        country, country_tok = None, None
        for t in desc_tokens:
            if t in lookups["country_lookup"]:
                country, country_tok = lookups["country_lookup"][t], t
                break
        if country is None:
            for t in header_tokens:
                if t in lookups["country_lookup"]:
                    country = lookups["country_lookup"][t]
                    break
        if country is None:
            country = sticky_country

        fruit, fruit_tok = None, None
        for t in desc_tokens:
            if t in lookups["fruit_lookup"]:
                fruit, fruit_tok = lookups["fruit_lookup"][t], t
                break
        if fruit is None:
            for t in header_tokens:
                if t in lookups["fruit_lookup"]:
                    fruit = lookups["fruit_lookup"][t]
                    break
        if fruit is None:
            fruit = sticky_fruit

        if country:
            sticky_country = country
        if fruit:
            sticky_fruit = fruit

        if not country or not fruit:
            flagged.append(make_flag(rec, "no_country_or_fruit_match",
                                      "Could not determine country and/or fruit from description or headers"))
            continue

        remaining = [t for t in desc_tokens if t != country_tok and t != fruit_tok]

        candidates = lookups["by_type_class"].get((fruit, country), [])
        if not candidates:
            flagged.append(make_flag(rec, "no_category_match",
                                      "No item-master rows exist for this fruit/country combination",
                                      fruit=fruit, country=country))
            continue

        variety_matched = None
        consumed = set()

        remaining_set = set(remaining)
        category_groups = defaultdict(list)
        for c in candidates:
            if c["ItemCategory"]:
                category_groups[c["ItemCategory"]].append(c)
        best_cat, best_consumed, best_score = None, set(), 0
        for cat in category_groups:
            cat_tokens = tokenize(cat)
            if not cat_tokens:
                continue
            matched = set()
            ok = True
            for ct in cat_tokens:
                hit = next((rt for rt in remaining_set
                            if rt == ct or rt.startswith(ct) or ct.startswith(rt)), None)
                if hit is None:
                    ok = False
                    break
                matched.add(hit)
            if ok and len(matched) > best_score:
                best_cat, best_consumed, best_score = cat, matched, len(matched)
        if best_cat:
            variety_matched, consumed = best_cat, best_consumed

        if variety_matched is None and fruit == "Apples" and "SMITH" in remaining:
            gs = [c for c in candidates if (c["ItemCategory"] or "").upper() == "GRANNY SMITH"]
            if gs:
                variety_matched = "Granny Smith"
                consumed = {"SMITH"}

        if variety_matched is None and len(category_groups) == 1:
            # Only one ItemCategory exists for this fruit/country combo at all, so
            # there's no real ambiguity even though the description names no variety.
            (only_cat,) = category_groups.keys()
            variety_matched = only_cat
            consumed = set()

        if variety_matched is None:
            # "Nil" is a real ItemCategory value (generic/no-specific-variety produce).
            # Fall back to it only when nothing more specific (e.g. "Organic") matched.
            nil_cats = [c for c in candidates if (c["ItemCategory"] or "").strip().upper() == "NIL"]
            if nil_cats:
                variety_matched = nil_cats[0]["ItemCategory"]
                consumed = set()

        if variety_matched is None:
            flagged.append(make_flag(rec, "no_category_match",
                                      "Description's variety words don't match any known ItemCategory",
                                      fruit=fruit, country=country,
                                      unmatched=" ".join(remaining)))
            continue

        remaining_after_variety = [t for t in remaining if t not in consumed]

        variety_candidates = [c for c in candidates if c["ItemCategory"] == variety_matched]

        size_num, size_str = parse_size(rec.get("size_raw"))
        if size_num is None:
            flagged.append(make_flag(rec, "size_not_resolved",
                                      "SIZE column value isn't a recognizable number",
                                      fruit=fruit, country=country,
                                      unmatched=" ".join(remaining_after_variety),
                                      candidates=variety_candidates))
            continue

        size_candidates = [c for c in variety_candidates if size_eq(c["Size"], size_num)]
        if not size_candidates:
            flagged.append(make_flag(rec, "size_not_resolved",
                                      "No item-master row at this size for the matched variety",
                                      fruit=fruit, country=country,
                                      unmatched=" ".join(remaining_after_variety),
                                      candidates=variety_candidates))
            continue

        grade_tokens = [t for t in remaining_after_variety if t in GRADE_WORDS]
        leftover = [t for t in remaining_after_variety if t not in IGNORE_WORDS and not t.isdigit()]
        distinct_codes = {c["ItemCode"] for c in size_candidates}

        if len(distinct_codes) > 1 and grade_tokens:
            # Grade (e.g. HG/SG) sometimes distinguishes SKUs that are otherwise
            # identical on fruit/country/variety/size/brand — use it before falling
            # back to brand matching or a duplicate-SKU auto-pick.
            grade_narrowed = [c for c in size_candidates if c["Grade"] in grade_tokens]
            distinct_narrowed = {c["ItemCode"] for c in grade_narrowed}
            if distinct_narrowed and len(distinct_narrowed) < len(distinct_codes):
                size_candidates = grade_narrowed
                distinct_codes = distinct_narrowed

        chosen = None
        verify_note = None

        if len(distinct_codes) == 1:
            chosen = size_candidates[0]
            chosen_brand = chosen["ItemBrand"].upper()
            mismatched_known_brand = [t for t in leftover if t in lookups["brand_set"] and t != chosen_brand]
            if mismatched_known_brand:
                flagged.append(make_flag(
                    rec, "unresolved_brand",
                    f"Description mentions brand '{' '.join(mismatched_known_brand)}' but the only "
                    f"item-master SKU at this size is brand '{chosen_brand}' — likely a missing SKU, "
                    "not a text-parsing issue",
                    fruit=fruit, country=country, unmatched=" ".join(leftover),
                    candidates=size_candidates))
                continue
            extra = [t for t in leftover if t != chosen_brand]
            if extra:
                verify_note = f"unrecognized text '{' '.join(extra)}' ignored"
        else:
            if leftover:
                brand_matches = [c for c in size_candidates if c["ItemBrand"].upper() in leftover]
            else:
                brand_matches = [c for c in size_candidates if c["ItemBrand"] == "Nil"]
            distinct_matched = {c["ItemCode"] for c in brand_matches}
            if len(distinct_matched) == 1:
                chosen = brand_matches[0]
                extra = [t for t in leftover if t != chosen["ItemBrand"].upper()]
                if extra:
                    verify_note = f"unrecognized text '{' '.join(extra)}' ignored"
            elif len(distinct_matched) > 1:
                sorted_codes = sorted(distinct_matched)
                chosen = next(c for c in brand_matches if c["ItemCode"] == sorted_codes[0])
                verify_note = f"duplicate SKU match ({'; '.join(sorted_codes)}), auto-picked {sorted_codes[0]}"
                flagged.append(make_flag(rec, "duplicate_sku_autopicked", verify_note,
                                          fruit=fruit, country=country,
                                          unmatched=" ".join(leftover),
                                          candidates=brand_matches))
            else:
                flagged.append(make_flag(rec, "unresolved_brand",
                                          f"Leftover text '{' '.join(leftover)}' doesn't match any known brand",
                                          fruit=fruit, country=country,
                                          unmatched=" ".join(leftover),
                                          candidates=size_candidates))
                continue

        duty_val = parse_duty(rec.get("duty_raw"))
        price_val = parse_price(rec.get("price_raw"))
        date_val = parse_date(rec.get("date_raw"))
        if duty_val is None or price_val is None or date_val is None:
            bad = []
            if duty_val is None:
                bad.append(f"duty={rec.get('duty_raw')!r}")
            if price_val is None:
                bad.append(f"price={rec.get('price_raw')!r}")
            if date_val is None:
                bad.append(f"date={rec.get('date_raw')!r}")
            flagged.append(make_flag(rec, "unparseable_field",
                                      "Could not parse: " + ", ".join(bad),
                                      fruit=fruit, country=country))
            continue

        currency = str(rec.get("currency") or "USD").strip().upper() or "USD"
        kg_display = str(rec["kg_raw"]).strip() if rec.get("kg_raw") is not None else ""
        detail_desc = f"{rec['desc_text']} {kg_display} {size_str}s".strip()
        remarks = str(rec["remark"]).strip() if rec.get("remark") else ""
        if verify_note:
            note = f"[VERIFY: {verify_note}]"
            remarks = f"{remarks} {note}".strip() if remarks else note

        resolved.append({
            "container": str(rec["container"]).strip(),
            "date": date_val,
            "item_code": chosen["ItemCode"],
            "detail_description": detail_desc,
            "qty": int(qty_num),
            "duty": f"RM {duty_val:.2f}",
            "price": f"{currency} {price_val:.2f}",
            "remarks": remarks,
            "source_row": rec["source_row"],
        })

    return resolved, flagged


def group_by_container(resolved):
    order = []
    groups = defaultdict(list)
    for rec in resolved:
        cont = rec["container"]
        if cont not in groups:
            order.append(cont)
        groups[cont].append(rec)
    return order, groups


def write_output(resolved, flagged, out_path):
    order, groups = group_by_container(resolved)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "StockReceive"
    ws.append(TEMPLATE_COLUMNS)

    today = datetime.datetime.now()
    date_conflicts = []

    for n, cont in enumerate(order, start=1):
        rows = groups[cont]
        dates = {r["date"] for r in rows}
        arrival_date = rows[0]["date"]
        if len(dates) > 1:
            date_conflicts.append((n, cont, sorted(dates)))

        for idx, r in enumerate(rows):
            row = [None] * len(TEMPLATE_COLUMNS)
            if idx == 0:
                row[TEMPLATE_COLUMNS.index("DocNo")] = f"Doc{n}"
                row[TEMPLATE_COLUMNS.index("DocDate")] = today
                row[TEMPLATE_COLUMNS.index("Description")] = "OPENING BALANCE"
                row[TEMPLATE_COLUMNS.index("UDF_RCV_MArrivalDate")] = arrival_date
            row[TEMPLATE_COLUMNS.index("ItemCode")] = r["item_code"]
            row[TEMPLATE_COLUMNS.index("DetailDescription")] = r["detail_description"]
            row[TEMPLATE_COLUMNS.index("Location")] = r["container"]
            row[TEMPLATE_COLUMNS.index("Qty")] = r["qty"]
            row[TEMPLATE_COLUMNS.index("UOM")] = "CTN"
            row[TEMPLATE_COLUMNS.index("UDF_RCVDtl_DnChar")] = r["duty"]
            row[TEMPLATE_COLUMNS.index("UDF_RCVDtl_Remarks")] = r["remarks"]
            row[TEMPLATE_COLUMNS.index("UDF_RCVDtl_STPrice")] = r["price"]
            ws.append(row)

    ws2 = wb.create_sheet("ReviewNeeded")
    ws2.append(REVIEW_COLUMNS)
    for fl in flagged:
        ws2.append([fl.get(col) for col in REVIEW_COLUMNS])
    for n, cont, dates in date_conflicts:
        ws2.append([
            f"Doc{n}", cont, None, None, None, "date_conflict_in_container",
            f"Container has multiple STOCK IN DATE values: {[d.strftime('%d/%m/%y') for d in dates]}; "
            f"used {dates[0].strftime('%d/%m/%y')} as the container arrival date",
            None, None, None, None, None,
        ])

    wb.save(out_path)
    return len(order), date_conflicts


def main():
    lookups = load_lookups()
    records = parse_gm_sheet(f"{BASE}/GM_stock_list.xlsx")
    resolved, flagged = resolve_records(records, lookups)

    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = f"{BASE}/GM_import_stock_receive_{ts}.xlsx"
    n_containers, date_conflicts = write_output(resolved, flagged, out_path)

    def _qty_positive(v):
        try:
            return v is not None and float(v) > 0
        except (TypeError, ValueError):
            return False

    qty_dropped = sum(1 for rec in records if not _qty_positive(rec.get("qty")))
    print(f"Rows parsed (pre qty filter): {len(records)}")
    print(f"Rows dropped for qty<=0: {qty_dropped}")
    print(f"Resolved rows (in StockReceive): {len(resolved)}")
    print(f"Flagged rows (in ReviewNeeded, excluded from main sheet unless dup-SKU): {len(flagged)}")
    print(f"Containers written: {n_containers}")
    print(f"Containers with conflicting arrival dates: {len(date_conflicts)}")
    print(f"Output written to: {out_path}")


if __name__ == "__main__":
    main()
